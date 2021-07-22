import openpyxl
import tkinter
from PIL import Image, ImageDraw, ImageFont

class rotto:
    def __init__(self,filename):
        self.rotoList=[]
        
        self.rotodata= openpyxl.load_workbook(f"./{filename}")
        self.rotoSheet=self.rotodata[filename.split('.')[0]]
        #print(self.rotoSheet.cell(3,5).value)
        #for self.t in self.rotoSheet.iter_rows(min_row=4,max_row=6, min_col=4, max_col=4):
         #   for self.col in self.t:
          #      print(self.col.value)
        self.i =4
        while True:
            self.rL=[]
            for self.r in self.rotoSheet[self.i]:
                self.rL.append(self.r.value)
            if self.rL[1] == None or self.rL[3] == None:
                print(f"데이터 끝 : {self.i}")
                
                break
            else:
                self.rotoList.append(self.rL)
                self.i+=1
        
    def printLottoList(self):
        return self.rotoList
            
       # print(self.rotoList)

#좌우 6씩간격
class rotoCanvas:
    def __init__(self):
        self.w = 600
        self.h= 1000

        #self.fnt= ImageFont.truetype("./fonts/NanumGothicBold.ttf", 50, encoding="UTF-8")
        self.fntN= ImageFont.truetype("./fonts/NanumGothicBold.ttf", 30, encoding="UTF-8")

        self.im = Image.new("RGBA", (self.w,self.h), (255,255,255))
        self.draw = ImageDraw.Draw(self.im)
        #self.draw.text((300,100),'999회',font=self.fnt, fill="black",align='center')
        
        self.draw.rectangle(((2,2),(600,200)), fill=(255,0,0))

        self.draw.rectangle(((3,10),(10,20)), outline=(255, 0, 0), width=2)
        #self.wt, self.ht = self.draw.textsize('999회',font=self.fnt)
        #self.draw.text(((self.w-self.wt)/2,80),'999회',font=self.fnt, fill="white",align='center')

        self.rx=14
        self.ry=230
        for self.j in range(0,7):
            self.rx=14

            for self.i in range(1,8):
                self.draw.rectangle(((self.rx,self.ry),(self.rx+57,self.ry+70)), outline=(255,0,0), width=1)
                self.draw.text((self.rx+13,self.ry+20),str(self.j*7+self.i),font=self.fntN, fill="red")
                self.rx+=57+28
                if self.i+self.j*7 == 45:
                    break
            self.ry+=70+40
            
        self.im.save("./original.png")
        #self.imb=self.im
    def printText(self,string,w,h,size):
        fnt= ImageFont.truetype("./fonts/NanumGothicBold.ttf", size, encoding="UTF-8")
        
        wt, ht = self.draw.textsize(string,font=fnt)
        self.draw.text(((w-wt)/2,80),string,font=fnt, fill="white",align='center')


    def makerottoImage(self,rottoList):
        self.image = Image.open("original.png")
        self.draw = ImageDraw.Draw(self.image)

        self.fnt= ImageFont.truetype("./fonts/NanumGothicBold.ttf", 20, encoding="UTF-8")

        self.im2=self.im
        #self.times=str(rottoList[1])+" 회"
        #self.wt, self.ht = self.draw.textsize(self.times,font=self.fnt)
        #self.draw.text(((self.w-self.wt)/2,80),self.times,font=self.fnt, fill="white",align='center')
        self.printText("제\t\t"+str(rottoList[1])+" 회",self.w,self.h,50)
        #추첨일
        self.draw.text((500,170),rottoList[2][2:],font=self.fnt,fill="white")

        for self.i in range(0,5):

            self.N=rottoList[self.i+3]-1
            self.N2=rottoList[self.i+4]-1

           # self.draw.line(((15+14+13+(85*(self.N%7))), (25+230+20+(self.N//7*110)),(15+14+13+(85*(self.N2%7))) , (25+230+20+(self.N2//7*110))), fill="skyblue", width=3)
            self.draw.line(((15+14+13+(85*(self.N%7))), (25+230+20+(self.N//7*110)),(15+14+13+(85*(self.N2%7))) , (25+230+20+(self.N2//7*110))), fill="skyblue", width=3)
            #self.draw.point(((15+14+13+(85*(self.N%7))), (25+230+20+(self.N//7*110))),"skyblue")
            self.rx, self.ry =(15+14+13+(85*(self.N%7))), (25+230+20+(self.N//7*110)) 
            self.rr=20
            self.draw.ellipse([(self.rx-self.rr,self.ry-self.rr),(self.rx+self.rr,self.ry+self.rr)], fill=(135,206,235), outline="skyblue")
        
        self.rx, self.ry =(15+14+13+(85*(self.N2%7))), (25+230+20+(self.N2//7*110)) 
        self.draw.ellipse([(self.rx-self.rr,self.ry-self.rr),(self.rx+self.rr,self.ry+self.rr)], fill=(135,206,235), outline="skyblue")
        self.rx, self.ry =(15+14+13+(85*((rottoList[9]-1)%7))), (25+230+20+((rottoList[9]-1)//7*110)) 
        #보너스 볼 번호 표시
        self.draw.ellipse([(self.rx-self.rr,self.ry-self.rr),(self.rx+self.rr,self.ry+self.rr)], fill=(211,211,211))

        #핑크
        #self.draw.ellipse([(self.rx-self.rr,self.ry-self.rr),(self.rx+self.rr,self.ry+self.rr)], fill=(255,182,193), outline="skyblue")




        self.rx=14
        self.ry=230
        for self.j in range(0,7):
            self.rx=14

            for self.i in range(1,8):
                self.draw.text((self.rx+13,self.ry+20),str(self.j*7+self.i),font=self.fntN, fill="red")
                self.rx+=57+28
                if self.i+self.j*7 == 45:
                    break
            self.ry+=70+40



        self.image.save(f"./images/{str(rottoList[1]).zfill(4)}회.png")
        
        
        

if __name__ =="__main__":
    
    a= input("파일 이름을 입력하세요: ")
    rotto1 = rotto(a)
    rottoList=rotto1.printLottoList()
    

    rotoCanvas1=rotoCanvas()
    for rl in rottoList:
        rotoCanvas1.makerottoImage(rl)
    

    #rotoCanvas1=rotoCanvas()
    #rotoCanvas1.makerottoImage([2021,969,"2021.07.08",3,4,5,6,7,8,9])
    
